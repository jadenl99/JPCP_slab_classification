# Developed by: Alvaro Bustindui Gonzalez

from PyQt5.uic import loadUi
import sys, os
import pandas as pd
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QStyle, QGridLayout, QPushButton, QVBoxLayout, QLabel, QGraphicsPixmapItem, QGraphicsScene
from PyQt5.QtGui import QPalette, QColor, QPixmap, QIcon
from PyQt5 import QtCore, QtGui, QtWidgets
import openpyxl
from os.path import exists
from PIL import Image

class MainWindow(QMainWindow):
    os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
    def __init__(self):

        self.application_path = ""
        if getattr(sys, 'frozen', False):
            self.application_path = os.path.dirname(sys.executable)
        elif __file__:
            self.application_path = os.path.dirname(__file__)

        self.datainput()

        super(MainWindow, self).__init__()
        loadUi(self.application_path + "/" + "mainapp.ui", self)
        self.showMaximized()
        screen_resolution = app.desktop().screenGeometry()
        width, height = screen_resolution.width(), screen_resolution.height()
        if width > 1930:
            width = 1930

        if height < 900:
            self.scrollArea.setFixedHeight(height - 80)

        self.scrollArea.setFixedWidth(width - 25)

        self.show()

        self.setup()

        self.nextButton.clicked.connect(self.next)
        self.backButton.clicked.connect(self.back)
        self.saveButton.clicked.connect(self.save)
        self.slabNo.returnPressed.connect(self.imgNo)

        self.buttonGroup1.buttonClicked.connect(self.click)
        self.yr1_label.clicked.connect(lambda: self.openimg(self.yr1_label))
        self.buttonGroup2.buttonClicked.connect(self.click)
        self.yr2_label.clicked.connect(lambda: self.openimg(self.yr2_label))
        if(self.yr3 != None):
            self.buttonGroup3.buttonClicked.connect(self.click)
            self.yr3_label.clicked.connect(lambda: self.openimg(self.yr3_label))
        if(self.yr4 != None):
            self.buttonGroup4.buttonClicked.connect(self.click)
            self.yr4_label.clicked.connect(lambda: self.openimg(self.yr4_label))
        if(self.yr5 != None):
            self.buttonGroup5.buttonClicked.connect(self.click)
            self.yr5_label.clicked.connect(lambda: self.openimg(self.yr5_label))
        if(self.yr6 != None):
            self.buttonGroup6.buttonClicked.connect(self.click)
            self.yr6_label.clicked.connect(lambda: self.openimg(self.yr6_label))

    def datainput(self):
        data_ready = False
        self.allyears = []
        while data_ready == False:
            try:
                actual_input = input("Enter years from least to greatest, separate them with a comma (ex. 2014, 2015, 2017):\n")
                
                if actual_input == "quit":
                    self.allyears = "quit"
                    raise Exception

                years_input = actual_input.split(", ")
            except: 
                if self.allyears == "quit":
                    sys.exit(0)
                continue

            self.yr1 = years_input[0]
            self.allyears.append(self.yr1)
            if(exists(self.application_path + "/" + self.yr1) == False):
                continue

            try:
                self.yr2 = years_input[1]
                self.allyears.append(self.yr2)
                if(exists(self.application_path + "/" + self.yr2) == False):
                    continue
            except: 
                continue

            try:
                self.yr3 = years_input[2]
                self.allyears.append(self.yr3)
                if(exists(self.application_path + "/" + self.yr3) == False):
                    continue
            except: 
                self.yr3=None
            
            try:
                self.yr4 = years_input[3]
                self.allyears.append(self.yr4)
                if(exists(self.application_path + "/" + self.yr4) == False):
                    continue
            except: 
                self.yr4=None
            
            try:
                self.yr5 = years_input[4]
                self.allyears.append(self.yr5)
                if(exists(self.application_path + "/" + self.yr5) == False):
                    continue
            except: 
                self.yr5=None
            
            try:
                self.yr6 = years_input[5]
                self.allyears.append(self.yr6)
                if(exists(self.application_path + "/" + self.yr6) == False):
                    continue
            except: 
                self.yr6=None
            
            data_ready = True
        
        ready = False
        while ready == False :
            try:
                inputpath = input("Enter the name of your datafile, make sure it is saved in the same directoy as where you are running this code (ex. data.xlsx):\n")
                
                if inputpath == "quit":
                    self.allyears = "quit"
                    raise Exception

                self.pathData = self.application_path + "/" + inputpath
                if(exists(self.pathData) == False):
                    continue

                for y in self.allyears:
                    try:
                        attempt = pd.read_excel(self.pathData, sheet_name = y, index_col=None)
                        try:
                            attemptcol = attempt[["Slab ID", "Slab State", "Slab State 2", "Comments", "Replaced"]]
                        except:
                            print("Missing columns in one of multiple sheets.")
                            break
                    except:
                        print("Missing sheet for one or multiple years.")
                        break
                    ready = True
            except:
                if self.allyears == "quit":
                    sys.exit(0)
        
        print("Opening application...")
    
    def newcheckhist(self):
        for y in self.alldata:
            index = self.slab + 1
            if (id(y["Results"]) != id(self.resultsyr1)):
                try:
                    index = int(y["Matching Dic"][self.slab + 1][0]["Slab ID"])
                except: 
                    continue

            try:
                y["Buttons"][y["Results"][index]["Slab State"]].setChecked(True)  
                y["Clicked"].append(y["Results"][index]["Slab State"])
            except:
                pass
            try:
                y["Buttons"][y["Results"][index]["Slab State 2"]].setChecked(True) 
                y["Buttons"][y["Results"][index]["Slab State 2"]].setIcon(self.style().standardIcon(QStyle.SP_DialogNoButton))
                y["Clicked"].append(y["Results"][index]["Slab State 2"])
            except: 
                pass

            try:
                y["Buttons"]["Comments"].setText(y["Results"][index]["Comments"])
            except: 
                pass

            try:
                if(int(y["Results"][index]["Replaced"]) == int(y["Year"])):
                    y["Buttons"]["Replaced"].setChecked(True)
            except: 
                pass

    def click(self, button):
        for y in self.alldata: 
            if button in y["Button Group"].buttons():
                if(button.text() in y["Clicked"]):
                    y["Clicked"].remove(button.text())
                    button.setChecked(False)    
                    button.setIcon(QIcon())
                    try:
                        y["Buttons"][y["Clicked"][0]].setIcon(QIcon())
                    except:
                        pass
                elif (len(y["Clicked"]) == 0):
                    y["Clicked"].append(button.text())
                    return
                elif (len(y["Clicked"]) == 1):
                    y["Clicked"].append(button.text())
                    y["Buttons"][button.text()].setIcon(self.style().standardIcon(QStyle.SP_DialogNoButton))
                    return
                
                if(len(y["Clicked"]) == 2):
                    button.setChecked(False)
    
    def openimg(self, label):
        yr = label.text()
        for y in self.alldata:
            if(yr == y["Year"]):
                try:
                    image = Image.open(y["Path to Img"])
                    image.show()
                except:
                    pass


    
    def cleanup(self):
        for y in self.alldata: 
            try: 
                y["Buttons"][y["Clicked"][0]].setChecked(False)  
            except:
                pass

            try: 
                y["Buttons"][y["Clicked"][1]].setIcon(QIcon())
                y["Buttons"][y["Clicked"][1]].setChecked(False)  
            except:
                pass

            y["Buttons"]["Replaced"].setChecked(False)

            y["Buttons"]["Comments"].setText("Comments...")
            y["Buttons"]["Comments"].clearFocus()

            y["Clicked"] = []


    def record(self):
        for y in self.alldata: 
            index = self.slab
            if (index >= list(self.resultsyr1.keys())[0] and id(y["Results"]) != id(self.resultsyr1)):
                try: 
                    index = int(y["Matching Dic"][self.slab][self.index_counter]["Slab ID"])
                except: 
                    continue
                

            try:
                y["Results"][index]["Slab State"] = y["Clicked"][0]
            except: 
                pass
                
            try:
                y["Results"][index]["Slab State 2"] = y["Clicked"][1]
            except: 
                pass

            try:
                if(y["Buttons"]["Comments"].text() != "Comments..."):
                    y["Results"][index]["Comments"] = y["Buttons"]["Comments"].text()
            except: 
                pass

            try:
                if(y["Buttons"]["Replaced"].isChecked()):
                    y["Results"][index]["Replaced"] = y["Year"]
            except: 
                pass
        

    def next(self, fromInside = False):
        if (fromInside == False):
            self.record()
        self.cleanup()
        self.newcheckhist()

        if (self.index - 1 == self.index_counter):
            self.slab += 1
            self.index = 1
            self.index_counter = 0

            self.slabNo.setText(str(self.slab))
            
            for y in self.alldata:
                neximg = None 
                if (id(y["Results"]) == id(self.resultsyr1) and self.slab in y["Results"]): 
                    nextimg = self.path1 + str(self.slab) + ".png"
                    y["Path to Img"] = nextimg
                    y["Img"].setPixmap(QPixmap(nextimg).scaled(300, 500, 1, 0))
                elif self.slab in y["Matching Dic"]:
                    self.index = max(self.index, len(y["Matching Dic"][self.slab]))
                    nextimg = y["Path"] + str(int(y["Matching Dic"][self.slab][self.index_counter]["Slab ID"])) + ".png"
                    y["Path to Img"] = nextimg
                    y["Img"].setPixmap(QPixmap(nextimg).scaled(300, 500, 1, 0))
                else: 
                    y["Img"].setText("NO IMAGE") 
        
        else: 
            self.index_counter += 1
            for y in self.alldata:
                nextimg = None
                try: 
                    nextimg = y["Path"] + str(int(y["Matching Dic"][self.slab][self.index_counter]["Slab ID"])) + ".png"
                    y["Path to Img"] = nextimg
                    y["Img"].setPixmap(QPixmap(nextimg).scaled(300, 500, 1, 0))
                except: 
                    y["Img"].setText("NO IMAGE") 

    
    def back(self):
        if(self.slab > list(self.resultsyr1.keys())[0]):
            self.slab -= 2
            self.index = 1
            self.index_counter = 0
            self.next(fromInside = True)
    
    def imgNo(self):
        self.record()
        if(int(self.slabNo.text()) in self.resultsyr1.keys()):
            self.slab = int(self.slabNo.text()) - 1
            self.index = 1
            self.index_counter = 0
            self.slabNo.clearFocus()
            self.next(fromInside = True)
    
    def setup(self):
        #self.yr1 = str(2015) # base year
        #self.yr2 = str(2016)
        #self.yr3 = str(2018)

        self.states = ["L1", "L2", "T1", "T2", "CC", "SS", "NC", "Error"]

        try:
            self.path1 = self.application_path + "/" + self.yr1 + "/Slabs/"
            self.nxtimg1 = ""
            self.path2 = self.application_path + "/" + self.yr2 + "/Slabs/"
            self.nxtimg2 = ""
            self.path3 = self.application_path + "/" + self.yr3 + "/Slabs/"
            self.nxtimg3 = ""
            self.path4 = self.application_path + "/" + self.yr4 + "/Slabs/"
            self.nxtimg4 = ""
            self.path5 = self.application_path + "/" + self.yr5 + "/Slabs/"
            self.nxtimg5 = ""
            self.path6 = self.application_path + "/" + self.yr6 + "/Slabs/"
            self.nxtimg6 = ""
        except:
            None
        
        #self.pathData = "inputforapp3.xlsx"

        try:
            self.exceldatayr1 = pd.read_excel(self.pathData, sheet_name = self.yr1, index_col=None)
            self.exceldatayr2 = pd.read_excel(self.pathData, sheet_name = self.yr2, index_col=None)
            self.exceldatayr3 = pd.read_excel(self.pathData, sheet_name = self.yr3, index_col=None)
            self.exceldatayr4 = pd.read_excel(self.pathData, sheet_name = self.yr4, index_col=None)
            self.exceldatayr5 = pd.read_excel(self.pathData, sheet_name = self.yr5, index_col=None)
            self.exceldatayr6 = pd.read_excel(self.pathData, sheet_name = self.yr6, index_col=None)
        except:
            None


        try:
            self.dictyr2 =  {k: g.to_dict(orient='records') for k, g in self.exceldatayr2[["Matching Slab ID", "Slab ID"]].set_index("Matching Slab ID").groupby(level=0)}
            self.dictyr3 =  {k: g.to_dict(orient='records') for k, g in self.exceldatayr3[["Matching Slab ID", "Slab ID"]].set_index("Matching Slab ID").groupby(level=0)}
            self.dictyr4 =  {k: g.to_dict(orient='records') for k, g in self.exceldatayr4[["Matching Slab ID", "Slab ID"]].set_index("Matching Slab ID").groupby(level=0)}
            self.dictyr5 =  {k: g.to_dict(orient='records') for k, g in self.exceldatayr5[["Matching Slab ID", "Slab ID"]].set_index("Matching Slab ID").groupby(level=0)}
            self.dictyr6 =  {k: g.to_dict(orient='records') for k, g in self.exceldatayr6[["Matching Slab ID", "Slab ID"]].set_index("Matching Slab ID").groupby(level=0)}
        except:
            None

        try:
            self.buttonsyr1 = {"L1": self.l1yr1, "L2": self.l2yr1, "T1": self.t1yr1, "T2": self.t2yr1, "CC": self.ccyr1, "SS": self.ssyr1, "NC": self.ncyr1, "Error": self.eyr1, "Comments": self.commentsyr1, "Replaced": self.repyr1}
            self.buttonsyr2 = {"L1": self.l1yr2, "L2": self.l2yr2, "T1": self.t1yr2, "T2": self.t2yr2, "CC": self.ccyr2, "SS": self.ssyr2, "NC": self.ncyr2, "Error": self.eyr2, "Comments": self.commentsyr2, "Replaced": self.repyr2}
            self.buttonsyr3 = {"L1": self.l1yr3, "L2": self.l2yr3, "T1": self.t1yr3, "T2": self.t2yr3, "CC": self.ccyr3, "SS": self.ssyr3, "NC": self.ncyr3, "Error": self.eyr3, "Comments": self.commentsyr3, "Replaced": self.repyr3}
            self.buttonsyr4 = {"L1": self.l1yr4, "L2": self.l2yr4, "T1": self.t1yr4, "T2": self.t2yr4, "CC": self.ccyr4, "SS": self.ssyr4, "NC": self.ncyr4, "Error": self.eyr4, "Comments": self.commentsyr4, "Replaced": self.repyr4}
            self.buttonsyr5 = {"L1": self.l1yr5, "L2": self.l2yr5, "T1": self.t1yr5, "T2": self.t2yr5, "CC": self.ccyr5, "SS": self.ssyr5, "NC": self.ncyr5, "Error": self.eyr5, "Comments": self.commentsyr5, "Replaced": self.repyr5}
            self.buttonsyr6 = {"L1": self.l1yr6, "L2": self.l2yr6, "T1": self.t1yr6, "T2": self.t2yr6, "CC": self.ccyr6, "SS": self.ssyr6, "NC": self.ncyr6, "Error": self.eyr6, "Comments": self.commentsyr6, "Replaced": self.repyr6}
        except:
            None

        try:
            self.clickedyr1 = []
            self.clickedyr2 = []
            self.clickedyr3 = []
            self.clickedyr4 = []
            self.clickedyr5 = []
            self.clickedyr6 = []
        except:
            None

        try:
            self.resultsyr1 = self.exceldatayr1[["Slab ID", "Slab State", "Slab State 2", "Replaced", "Comments"]].set_index("Slab ID").to_dict(orient = "index")
            self.resultsyr2 = self.exceldatayr2[["Slab ID", "Slab State", "Slab State 2", "Replaced", "Comments"]].set_index("Slab ID").to_dict(orient = "index")
            self.resultsyr3 = self.exceldatayr3[["Slab ID", "Slab State", "Slab State 2", "Replaced", "Comments"]].set_index("Slab ID").to_dict(orient = "index")
            self.resultsyr4 = self.exceldatayr4[["Slab ID", "Slab State", "Slab State 2", "Replaced", "Comments"]].set_index("Slab ID").to_dict(orient = "index")
            self.resultsyr5 = self.exceldatayr5[["Slab ID", "Slab State", "Slab State 2", "Replaced", "Comments"]].set_index("Slab ID").to_dict(orient = "index")
            self.resultsyr6 = self.exceldatayr6[["Slab ID", "Slab State", "Slab State 2", "Replaced", "Comments"]].set_index("Slab ID").to_dict(orient = "index")
        except:
            None


        try:
            self.datayr1 = {"Year": self.yr1, "Data": self.exceldatayr1, "Buttons": self.buttonsyr1, "Clicked": self.clickedyr1, "Results": self.resultsyr1, "Button Group": self.buttonGroup1, "Path": self.path1, "Img": self.img1, "Path to Img": self.nxtimg1}
            self.datayr2 = {"Year": self.yr2, "Data": self.exceldatayr2, "Buttons": self.buttonsyr2, "Clicked": self.clickedyr2, "Results": self.resultsyr2, "Button Group": self.buttonGroup2, "Matching Dic": self.dictyr2, "Path": self.path2, "Img": self.img2, "Path to Img": self.nxtimg2}
            self.datayr3 = {"Year": self.yr3, "Data": self.exceldatayr3, "Buttons": self.buttonsyr3, "Clicked": self.clickedyr3, "Results": self.resultsyr3, "Button Group": self.buttonGroup3, "Matching Dic": self.dictyr3, "Path": self.path3, "Img": self.img3, "Path to Img": self.nxtimg3}
            self.datayr4 = {"Year": self.yr4, "Data": self.exceldatayr4, "Buttons": self.buttonsyr4, "Clicked": self.clickedyr4, "Results": self.resultsyr4, "Button Group": self.buttonGroup4, "Matching Dic": self.dictyr4, "Path": self.path4, "Img": self.img4, "Path to Img": self.nxtimg4}
            self.datayr5 = {"Year": self.yr5, "Data": self.exceldatayr5, "Buttons": self.buttonsyr5, "Clicked": self.clickedyr5, "Results": self.resultsyr5, "Button Group": self.buttonGroup5, "Matching Dic": self.dictyr5, "Path": self.path5, "Img": self.img5, "Path to Img": self.nxtimg5}
            self.datayr6 = {"Year": self.yr6, "Data": self.exceldatayr6, "Buttons": self.buttonsyr6, "Clicked": self.clickedyr6, "Results": self.resultsyr6, "Button Group": self.buttonGroup6, "Matching Dic": self.dictyr6, "Path": self.path6, "Img": self.img6, "Path to Img": self.nxtimg6}
        except:
            None

        self.alldata = [self.datayr1, self.datayr2]
        if(self.yr3 != None):
            self.alldata.append(self.datayr3)
        if(self.yr4 != None):
            self.alldata.append(self.datayr4)
        if(self.yr5 != None):
            self.alldata.append(self.datayr5)
        if(self.yr6 != None):
            self.alldata.append(self.datayr6)
            

        try:
            self.yr1_label.setText(self.yr1)
            self.yr2_label.setText(self.yr2)
            self.yr3_label.setText(self.yr3)
            self.yr4_label.setText(self.yr4)
            self.yr5_label.setText(self.yr5)
            self.yr6_label.setText(self.yr6)
        except:
            None


        self.numLabel.setText("/" + str(list(self.resultsyr1.keys())[-1]))

        self.slab = list(self.resultsyr1.keys())[0] - 1
        self.index = 1
        self.index_counter = 0

    def save(self):
        self.record()
        book = openpyxl.load_workbook(self.pathData)
        active_sheet = book.active

        for y in self.alldata:
            yr_df = pd.read_excel(self.pathData, sheet_name = y["Year"], index_col=None)
            c_ss = yr_df.columns.get_loc("Slab State") + 1
            c_ss2 = yr_df.columns.get_loc("Slab State 2") + 1
            c_comm = yr_df.columns.get_loc("Comments") + 1
            c_rep = yr_df.columns.get_loc("Replaced") + 1


            sheet = book.get_sheet_by_name(y["Year"])
            for i in y["Results"]:
                r = yr_df[yr_df['Slab ID'] == i].index[0] + 2

                cell = sheet.cell(row = r, column = c_ss)
                cell.value = y["Results"][i]["Slab State"]

                cell2 = sheet.cell(row = r, column = c_ss2)
                cell2.value = y["Results"][i]["Slab State 2"]

                cell3 = sheet.cell(row = r, column = c_comm)
                cell3.value = y["Results"][i]["Comments"]

                cell4 = sheet.cell(row = r, column = c_rep)
                cell4.value = y["Results"][i]["Replaced"]

        book.save(self.pathData)
            
app = QApplication(sys.argv)

UIWindow = MainWindow()
app.exec_()
